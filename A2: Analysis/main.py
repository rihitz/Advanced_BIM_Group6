from pathlib import Path
import ifcopenshell  
import pandas as pd  
import openpyxl

def loadFile():
    # Specify the name of the IFC model
    modelname = input("Enter the IFC model name (or press Enter for default 'LLYN - STRU'): ")
    
    # Set a default value if the user didn't provide input
    if not modelname:
        modelname = 'LLYN - STRU'
    
    try:
        dir_path = Path(__file__).parent
        model_url = Path.joinpath(dir_path, 'model', modelname).with_suffix('.ifc')
        model = ifcopenshell.open(model_url)
    except OSError:
        try:
            # If file not found, bpy import will produce an error, if script is run outside of Blender
            import bpy
            model_url = Path.joinpath(Path(bpy.context.space_data.text.filepath).parent, 'model', modelname).with_suffix('.ifc')
            model = ifcopenshell.open(model_url)
        except OSError:
            print(f'ERROR: please check your model folder : {model_url} does not exist')
    
    # Load the IFC model
    model = ifcopenshell.open(model_url)
    
    return model, modelname

def quantityTakeOff(element):
    # Retrieve IFC elements of the specified type
    str_elements = model.by_type(str(element))
    tot_elements = len(str_elements)
    
    # Create an empty pandas DataFrame with column titles
    columns = ['Floor', 'Volume, m3', 'Area, m2', 'Material']
    df = pd.DataFrame(columns=columns)

    # Check if the IFC model contains specified structural elements
    if len(str_elements) != 0:
       # Iterate through elements of the specified type
        for str_element in str_elements:
            if str_element.IsDefinedBy:
                definitions = str_element.IsDefinedBy
    
                is_load_bearing = False  # Initialize the load-bearing flag
    
                for definition in definitions:
                    # Check if the relationship is of type 'IfcRelDefinesByProperties'
                    if definition.is_a('IfcRelDefinesByProperties'):
                        property_definition = definition.RelatingPropertyDefinition
    
                        # Check if the element is load-bearing (assuming 'LoadBearing' property)
                        if property_definition.is_a('IfcPropertySet'):
                            for property in property_definition.HasProperties:
                                if property.Name == 'LoadBearing':
                                    # Check if the property value is True
                                    if property.NominalValue.wrappedValue == True:
                                        is_load_bearing = True  # Set the flag to True and break
                                        break  # Exit the property loop once load-bearing is confirmed
    
                # If the element is load-bearing, proceed to calculate quantities
                if is_load_bearing:
                    floor_name = str_element.ContainedInStructure[0].RelatingStructure.Name
                    
                    # Create a dictionary to store data for a DataFrame entry
                    element_properties = {
                        'Floor': floor_name,
                        'Volume': None,
                        'Area': None,
                        'Material': None
                    }
    
                    for definition in definitions:
                        # Check if the relationship is of type 'IfcRelDefinesByProperties'
                        if definition.is_a('IfcRelDefinesByProperties'):
                            property_definition = definition.RelatingPropertyDefinition
    
                            if property_definition.is_a('IfcElementQuantity'):
                                for quantity in property_definition.Quantities:
                                    
                                    # Check if the quantity is of type 'IfcQuantityArea'
                                    if quantity.is_a('IfcQuantityArea'):
                                        
                                        # Slab area, m2
                                        if quantity.Name == 'NetArea':
                                            element_properties['Area, m2'] = quantity.AreaValue
                                            
                                        # Wall surface area, m2
                                        elif quantity.Name == 'NetSideArea':
                                            element_properties['Area, m2'] = quantity.AreaValue
                                            
                                        # Beam surface area, m2
                                        elif quantity.Name == 'NetSurfaceArea':
                                            element_properties['Area, m2'] = quantity.AreaValue
                                            
                                    # Check if the quantity is of type 'IfcQuantityVolume' 
                                    elif quantity.is_a('IfcQuantityVolume'):
                                        
                                        # Element volume, m3
                                        if quantity.Name == 'NetVolume':
                                            element_properties['Volume, m3'] = quantity.VolumeValue
    
                    # for definition in definitions:
                    #     # Check if the relationship is of type 'IfcRelAssociatesMaterial'
                    #     if definition.is_a('IfcRelAssociatesMaterial'):
                    #         # This relationship associates materials with the element
                    #         # Iterate through related materials
                    #         for related_material in definition.RelatedObjects:
                    #             if related_material.is_a('IfcMaterial'):
                    #                 # Extract material name and add it to the list
                    #                 element_properties['Material'] = related_material.Name
    
                    # Append the data as a new row to the pandas DataFrame
                    df = pd.concat([df, pd.DataFrame([element_properties], columns=columns)], ignore_index=True)
                    
                    # Create a floorwise summary:
                    floorwise_sum = df.groupby('Floor')[['Volume, m3', 'Area, m2']].sum()
                    
                    # Calculate total quantities for the specified element type
                    total_volume = df['Volume, m3'].sum()
                    total_area = df['Area, m2'].sum()
                

    return df, tot_elements, floorwise_sum, total_volume, total_area, element

def report(floorwise_summary, tot_quantities, modelname):

    # Concatenate floorwise summary results
    summary = pd.concat(floorwise_summary, axis=1).round(2)
    
    # Reset the index to apply numerical indexing and make the "Floor" index an actual column
    summary.reset_index(level=0, inplace=True)
    summary.index += 1  # Add 1 to the index to start indexing from 1

    # Rename the columns to distinguish duplicate names
    summary.columns = [f'{col} {i}' for i, col in enumerate(summary.columns)]
    
    # Convert to DataFrame
    tot_quantities = pd.DataFrame([tot_quantities])
    
    # Load template and save a report file
    wb_temp = openpyxl.load_workbook('./input/rep_template.xlsx')
    wb_temp.save('./output/' + "report_" + modelname + '.xlsx')

    report_file_path = './output/' + "report_" + modelname + '.xlsx'
    wb_rep = openpyxl.load_workbook(report_file_path)

    # Create space for reported quantities while maintaining template layout
    for i in range(1,len(summary)): 
        wb_rep['Sheet1'].insert_rows(6) # number indicates insertion point
    
    # Add index column
    for i in range(1,len(summary)+1): 
        wb_rep['Sheet1'].cell(row=5+i, column=2).value = i 

    # Insert IFC file name
    wb_rep['Sheet1'].cell(row=3, column=3).value = modelname + ".ifc"
    wb_rep.save(report_file_path)
    
    # Create a Pandas ExcelWriter to append data
    with pd.ExcelWriter(report_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:

        # Write the DataFrame to the Excel file on 'Sheet1'
        summary.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=5, startcol=2)
        tot_quantities.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=5+len(summary), startcol=3)

    # Inform about the completion and storage location
    print("Report has been exported to an Excel file:", report_file_path)
    
    return 

# Load the IFC model
model, modelname = loadFile()  

# Output a written report within Python's console:  

# Collect floor names from the IFC model
floorNames = []

for entity in model.by_type('IfcBuildingStorey'):
    floorNames.append(entity.Name)

print('\nThere are', len(floorNames), 'floors in the model: \n')

# Print the names of each floor
for floor in floorNames:
    print(floor)

# List of structural element types to be reported
elements = ['IfcColumn', 'IfcWall', 'IfcBeam', 'IfcSlab']

# Initialize lists for result storage:
element_summary = []        # list of DataFrames for each element
floorwise_summary = []      # list of DataFrames for floorwise summaries
tot_quantities = []         # total quantities by element type

# Process each structural element type
for element in elements:
    element_quantities, tot_elements, floorwise_sum, total_volume, total_area , element = quantityTakeOff(element)
    
    # Add results to the lists
    element_summary.append(element_quantities)
    floorwise_summary.append(floorwise_sum)
    tot_quantities.append(total_volume)
    tot_quantities.append(total_area)
    
    # Print a report in Python's console
    print('________________________________________________________________')
    print('\nOut of', tot_elements, element, 'elements', len(element_quantities), 'are load bearing (structural).')
    print('\nFloorwise summary for', element,' elements: \n')
    print(floorwise_sum, '\n')
    print('Total material quantities for', element, 'elements:')
    print('Volume: ', round(total_volume,2), 'm3')
    print('Surface Area: ', round(total_area,2), 'm2')
    
# To access the full DataFrame of an element:
    # element_summary[0] - Columns
    # element_summary[1] - Walls
    # element_summary[2] - Beams
    # element_summary[3] - Slabs
    
    # (to be used in further processing for price calculations)


# Generate a spreadsheet report
while True: 
    prompt = input("\nDo you want to generate a spreadsheet report? (Enter 'y' for Yes, 'n' for No): ")
    
    if prompt == 'y':   
        report(floorwise_summary, tot_quantities, modelname)
        break
    elif prompt == 'n':
        print("Exiting without generating a report.")
        break
    else:
        print("Invalid input. Please enter 'y' or 'n'.")  





